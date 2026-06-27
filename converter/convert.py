#!/usr/bin/env python3
"""
Frei Lab Compound Database — Excel → JSON converter.

Usage:
    python convert.py <excel_file> <output_dir> --library-id <ID>

Supported library IDs:
    IrCpSB    — Ir Cp Schiff-Base combinatorial library (single sheet)
    TzLib     — Metal-Triazole combinatorial library (multi-sheet)
    NOSB      — N,O-Schiff Base Ru/Ir combinatorial library (multi-sheet)
    MnSB      — Manganese(I) Schiff-Base combinatorial library (single sheet)

Writes:
    <output_dir>/manifest.json
    <output_dir>/libraries/<ID>.json
"""

import sys
import json
import re
import argparse
from pathlib import Path

import openpyxl
from rdkit import Chem
from rdkit.Chem.Draw import rdMolDraw2D
from rdkit.Chem.inchi import MolToInchiKey

# ═══════════════════════════════════════════════════════════════════════════════
# Shared metadata — add a new entry here for each new library
# ═══════════════════════════════════════════════════════════════════════════════

LIBRARY_META = {
    "IrCpSB": {
        "title": "Ir Cp Schiff-Base Library",
        "description": (
            "1,440 iridium Cp Schiff-base compounds from a combinatorial library "
            "screened for antibacterial activity against S. aureus and E. coli, "
            "and for cytotoxicity against HEK293T cells."
        ),
        "metal": "Ir",
        "scaffold": "Cp Schiff-Base",
        "doi": "10.26434/chemrxiv.15003837/v1",
    },
    "TzLib": {
        "title": "Metal-Triazole Combinatorial Library",
        "description": (
            "864 metal-triazole compounds from two combinatorial series "
            "(Tz-4-P and Tz-1-MP) across multiple metal scaffolds "
            "(IrCN, Re(CO)₃, Mn(CO)₃, RuCy), screened for "
            "antibacterial activity against S. aureus and cytotoxicity "
            "against HEK293T cells."
        ),
        "metal": "Ir / Re / Mn / Ru",
        "scaffold": "Triazole",
        "doi": "10.1038/s41467-025-67341-z",
    },
    "NOSB": {
        "title": "N,O-Schiff Base Ru/Ir Library",
        "description": (
            "176 N,O-Schiff-base compounds from a combinatorial library across "
            "two half-sandwich metal scaffolds (Ir Cp*, Ru cymene), screened for "
            "antibacterial activity against S. aureus and E. coli, and "
            "cytotoxicity against HEK293T cells."
        ),
        "metal": "Ir / Ru",
        "scaffold": "N,O-Schiff Base",
        "doi": "10.26434/chemrxiv.15005024/v1",
    },
    "MnSB": {
        "title": "Manganese(I) Schiff-Base Library",
        "description": (
            "420 manganese(I) tricarbonyl Schiff-base compounds from a "
            "combinatorial library (6 axial ligands × 10 amines × 7 aldehydes), "
            "screened for antibacterial activity against MRSA."
        ),
        "metal": "Mn",
        "scaffold": "Schiff Base",
        "doi": "10.1039/D3SC05326A",
    },
    "IrCN_Click": {
        "title": "IrCN Click-Chemistry Library",
        "description": (
            "288 cyclometalated iridium(III) click-chemistry compounds from a "
            "combinatorial library (3 metal scaffolds × 24 amines × 4 alkynes), "
            "screened for antibacterial activity against S. aureus."
        ),
        "metal": "Ir",
        "scaffold": "IrCN Click",
        "doi": None,
    },
    "IrCN_Schiff": {
        "title": "IrCN Schiff-Base Library",
        "description": (
            "264 cyclometalated iridium(III) Schiff-base compounds from a "
            "combinatorial library (3 metal scaffolds × 8 aldehydes × 11 amines), "
            "screened for antibacterial activity against S. aureus."
        ),
        "metal": "Ir",
        "scaffold": "IrCN Schiff Base",
        "doi": None,
    },
}

# ═══════════════════════════════════════════════════════════════════════════════
# Shared helpers
# ═══════════════════════════════════════════════════════════════════════════════

def clean_val(v):
    """Float or None — treats '-', 'N.A.', empty string as missing."""
    if v is None or (isinstance(v, str) and v.strip() in ("-", "", "N.A.", "N/A", "NA")):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def smiles_to_svg(smiles, width=220, height=180):
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return None
    drawer = rdMolDraw2D.MolDraw2DSVG(width, height)
    drawer.drawOptions().addStereoAnnotation = True
    drawer.DrawMolecule(mol)
    drawer.FinishDrawing()
    return drawer.GetDrawingText()


def smiles_to_inchikey(smiles):
    mol = Chem.MolFromSmiles(smiles)
    return MolToInchiKey(mol) if mol else None


def render_building_blocks(bb_map):
    """
    bb_map: {position_key: {smiles_or_None: code}}
    Returns building_blocks dict ready for the JSON output.
    Positions whose SMILES are None get null svg/canonical_key.
    """
    building_blocks = {}
    for pos_key, entries in bb_map.items():
        building_blocks[pos_key] = []
        for smiles, code in sorted(entries.items(), key=lambda x: x[1]):
            if smiles:
                print(f"  {pos_key} {code} ...", end=" ", flush=True)
                svg = smiles_to_svg(smiles)
                ik  = smiles_to_inchikey(smiles)
                print("ok")
            else:
                svg, ik = None, None
            building_blocks[pos_key].append({
                "code": code,
                "smiles": smiles,
                "name": None,
                "canonical_key": ik,
                "svg": svg,
            })
    return building_blocks


def write_outputs(library, output_dir):
    """Write the library JSON and update manifest.json."""
    out     = Path(output_dir)
    lib_dir = out / "libraries"
    lib_dir.mkdir(parents=True, exist_ok=True)

    lib_path = lib_dir / f"{library['id']}.json"
    with open(lib_path, "w", encoding="utf-8") as f:
        json.dump(library, f, ensure_ascii=False, separators=(",", ":"))
    print(f"Wrote {lib_path}  ({lib_path.stat().st_size // 1024} KB)")

    manifest_path = out / "manifest.json"
    if manifest_path.exists():
        with open(manifest_path, encoding="utf-8") as f:
            manifest = json.load(f)
    else:
        manifest = {"libraries": []}

    meta      = LIBRARY_META[library["id"]]
    lib_id    = library["id"]
    entry = {
        "id":             lib_id,
        "title":          meta["title"],
        "description":    meta["description"],
        "metal":          meta["metal"],
        "scaffold":       meta["scaffold"],
        "doi":            meta["doi"],
        "compound_count": len(library["compounds"]),
        "position_count": len(library["positions"]),
    }
    existing = {lib["id"] for lib in manifest["libraries"]}
    if lib_id in existing:
        manifest["libraries"] = [
            entry if lib["id"] == lib_id else lib
            for lib in manifest["libraries"]
        ]
    else:
        manifest["libraries"].append(entry)

    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    print(f"Wrote {manifest_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# IrCpSB — Ir Cp Schiff-Base (single sheet, 3 SMILES columns, 4-rep assays)
# ═══════════════════════════════════════════════════════════════════════════════

IRCPSB_POSITIONS = [
    {"key": "Cp",    "label": "Cp Ligand", "col": 1},
    {"key": "Ald",   "label": "Aldehyde",  "col": 2},
    {"key": "Amine", "label": "Amine",     "col": 3},
]

IRCPSB_PROPERTIES = [
    {"key": "conversion", "label": "Conversion",          "unit": "%",   "role": "qc",      "group": None,            "col": 4,  "rep_cols": None},
    {"key": "rt_target",  "label": "RT (Target)",         "unit": "min", "role": "qc",      "group": None,            "col": 5,  "rep_cols": None},
    {"key": "rt_2plus",   "label": "RT (2+)",             "unit": "min", "role": "qc",      "group": None,            "col": 6,  "rep_cols": None},
    {"key": "sa_50",      "label": "S. aureus 50 µM",  "unit": "OD",  "role": "primary", "group": "Antibacterial", "col": 11, "rep_cols": [7, 8, 9, 10]},
    {"key": "sa_12",      "label": "S. aureus 12.5 µM","unit": "OD",  "role": "primary", "group": "Antibacterial", "col": 16, "rep_cols": [12, 13, 14, 15]},
    {"key": "ec_50",      "label": "E. coli 50 µM",   "unit": "OD",  "role": "primary", "group": "Antibacterial", "col": 21, "rep_cols": [17, 18, 19, 20]},
    {"key": "hek_50",     "label": "HEK293T 50 µM",   "unit": "%",   "role": "primary", "group": "Cytotoxicity",  "col": 26, "rep_cols": [22, 23, 24, 25]},
    {"key": "ratio",      "label": "Selectivity Ratio",   "unit": None,  "role": "derived", "group": None,            "col": 27, "rep_cols": None},
]


def _parse_ircpsb_id(cid):
    m = re.match(r"IrCp(\d+)([A-Z]+)(\d+)$", str(cid))
    if not m:
        return None
    return {"Cp": f"Cp{m.group(1)}", "Ald": m.group(2), "Amine": m.group(3)}


def convert_ircpsb(wb, library_id):
    ws        = wb.active
    data_rows = [r for r in ws.iter_rows(values_only=True) if r[0] and r[0] != "Compounds"]
    print(f"  {len(data_rows)} compound rows")

    # Collect unique SMILES per position
    bb_map = {p["key"]: {} for p in IRCPSB_POSITIONS}
    skipped = 0
    for row in data_rows:
        blocks = _parse_ircpsb_id(row[0])
        if not blocks:
            print(f"  Warning: cannot parse '{row[0]}' — skipping")
            skipped += 1
            continue
        for pos in IRCPSB_POSITIONS:
            smiles = row[pos["col"]]
            if smiles and smiles not in bb_map[pos["key"]]:
                bb_map[pos["key"]][smiles] = blocks[pos["key"]]
    if skipped:
        print(f"  {skipped} rows skipped")
    for pos in IRCPSB_POSITIONS:
        print(f"  {pos['key']}: {len(bb_map[pos['key']])} unique")

    print("Generating SVGs ...")
    building_blocks = render_building_blocks(bb_map)

    print("Building compound records ...")
    compounds = []
    for row in data_rows:
        blocks = _parse_ircpsb_id(row[0])
        if not blocks:
            continue
        props = {}
        for p in IRCPSB_PROPERTIES:
            avg = clean_val(row[p["col"]])
            if p["rep_cols"]:
                props[p["key"]] = {"avg": avg, "reps": [clean_val(row[i]) for i in p["rep_cols"]]}
            else:
                props[p["key"]] = avg
        compounds.append({"id": row[0], "blocks": blocks, "props": props})

    meta = LIBRARY_META[library_id]
    return {
        "id": library_id, "title": meta["title"], "description": meta["description"],
        "metal": meta["metal"], "scaffold": meta["scaffold"], "doi": meta["doi"],
        "positions":       [{"key": p["key"], "label": p["label"]} for p in IRCPSB_POSITIONS],
        "properties":      [{k: v for k, v in p.items() if k not in ("col", "rep_cols")} for p in IRCPSB_PROPERTIES],
        "building_blocks": building_blocks,
        "compounds":       compounds,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# TzLib — Metal-Triazole library (multi-sheet, concatenated SMILES, no reps)
# ═══════════════════════════════════════════════════════════════════════════════

TZ_POSITIONS = [
    {"key": "Scaffold", "label": "Metal Scaffold"},
    {"key": "Amine",    "label": "Amine"},
    {"key": "Alkyne",   "label": "Alkyne"},
]

TZ_PROPERTIES = [
    {"key": "peak_pct",  "label": "Conversion",         "unit": "%",   "role": "qc",       "group": None},
    {"key": "peak_norm", "label": "Conversion (norm.)", "unit": "%",   "role": "qc",       "group": None},
    {"key": "rt",        "label": "RT",                 "unit": "min", "role": "qc",       "group": None},
    {"key": "sdr",       "label": "SDR (S. aureus)",    "unit": "µM",  "role": "primary",  "group": "Antibacterial"},
    {"key": "mic",       "label": "MIC (S. aureus)",    "unit": "µM",  "role": "primary",  "group": "Antibacterial"},
    {"key": "tox_avg",   "label": "HEK293T growth",     "unit": "%",   "role": "primary",  "group": "Cytotoxicity"},
    {"key": "tox_sd",    "label": "HEK293T SD",         "unit": "%",   "role": "replicate","group": "Cytotoxicity"},
]

# Each entry describes one sheet: how to identify it, parse its compound IDs,
# and which columns hold each property.
TZ_SHEET_CONFIGS = [
    # ── Tz-4-P series (24 amines M1–M24 × 4 alkynes Y1–Y4) ──────────────────
    {
        "sheet": "Tz-4-P",
        "scaffold_code":  "Free_Tz4P",
        "scaffold_label": "Free triazole (Tz-4-P)",
        "amine_prefix": "M", "alkyne_prefix": "Y",
        "id_re": re.compile(r"^M(\d+)Y(\d+)$"),
        "cols": {"peak_pct": 2, "peak_norm": None, "rt": 3, "sdr": 4, "mic": 5, "tox_avg": 6, "tox_sd": 7},
    },
    {
        "sheet": "IrCN",
        "scaffold_code":  "IrCN",
        "scaffold_label": "IrCN",
        "amine_prefix": "M", "alkyne_prefix": "Y",
        "id_re": re.compile(r"^IrCN_M(\d+)Y(\d+)$"),
        "cols": {"peak_pct": 2, "peak_norm": None, "rt": 3, "sdr": 4, "mic": 5, "tox_avg": 6, "tox_sd": 7},
    },
    {
        "sheet": "Re(CO)3",
        "scaffold_code":  "Re_CO3",
        "scaffold_label": "Re(CO)₃",
        "amine_prefix": "M", "alkyne_prefix": "Y",
        "id_re": re.compile(r"^Re\(CO\)3M(\d+)Y(\d+)$"),
        "cols": {"peak_pct": 2, "peak_norm": None, "rt": 3, "sdr": 4, "mic": 5, "tox_avg": 6, "tox_sd": 7},
    },
    {
        "sheet": "Re(CO)3 solvent",
        "scaffold_code":  "Re_CO3_solv",
        "scaffold_label": "Re(CO)₃ (solvento)",
        "amine_prefix": "M", "alkyne_prefix": "Y",
        "id_re": re.compile(r"^Re\(CO\)3M(\d+)Y(\d+)solvent$"),
        "cols": {"peak_pct": 2, "peak_norm": None, "rt": 3, "sdr": 4, "mic": None, "tox_avg": None, "tox_sd": None},
    },
    {
        "sheet": "Mn(CO)3",
        "scaffold_code":  "Mn_CO3",
        "scaffold_label": "Mn(CO)₃",
        "amine_prefix": "M", "alkyne_prefix": "Y",
        "id_re": re.compile(r"^Mn\(CO\)3M(\d+)Y(\d+)$"),
        "cols": {"peak_pct": 2, "peak_norm": None, "rt": 3, "sdr": 4, "mic": 5, "tox_avg": 6, "tox_sd": 7},
    },
    {
        "sheet": "Mn(CO)3 solvent",
        "scaffold_code":  "Mn_CO3_solv",
        "scaffold_label": "Mn(CO)₃ (solvento)",
        "amine_prefix": "M", "alkyne_prefix": "Y",
        "id_re": re.compile(r"^Mn\(CO\)3M(\d+)Y(\d+)solvent$"),
        "cols": {"peak_pct": 2, "peak_norm": None, "rt": 3, "sdr": 4, "mic": None, "tox_avg": None, "tox_sd": None},
    },
    {
        "sheet": "RuCy(Tz-4-P)",
        "scaffold_code":  "RuCy_Tz4P",
        "scaffold_label": "RuCy (Tz-4-P)",
        "amine_prefix": "M", "alkyne_prefix": "Y",
        "id_re": re.compile(r"^RuCyM(\d+)Y(\d+)$"),
        "cols": {"peak_pct": 2, "peak_norm": 3, "rt": 4, "sdr": 5, "mic": 6, "tox_avg": None, "tox_sd": None},
    },
    # ── Tz-1-MP series (8 amines P1–P8 × 12 alkynes A1–A12) ─────────────────
    {
        "sheet": "Tz-1-MP",
        "scaffold_code":  "Free_Tz1MP",
        "scaffold_label": "Free triazole (Tz-1-MP)",
        "amine_prefix": "P", "alkyne_prefix": "A",
        "id_re": re.compile(r"^P(\d+)A(\d+)$"),
        "cols": {"peak_pct": 2, "peak_norm": None, "rt": 3, "sdr": 4, "mic": None, "tox_avg": 5, "tox_sd": 6},
    },
    {
        "sheet": "RuCy(Tz-1-MP) ",  # trailing space in workbook
        "scaffold_code":  "RuCy_Tz1MP",
        "scaffold_label": "RuCy (Tz-1-MP)",
        "amine_prefix": "P", "alkyne_prefix": "A",
        "id_re": re.compile(r"^RuCyP(\d+)A(\d+)$"),
        "cols": {"peak_pct": 2, "peak_norm": 3, "rt": 4, "sdr": 5, "mic": None, "tox_avg": 6, "tox_sd": 7},
    },
]


def convert_tzlib(wb, library_id):
    # bb_map: position_key → {smiles_or_None: code}
    bb_map = {"Scaffold": {}, "Amine": {}, "Alkyne": {}}
    all_compounds = []

    for cfg in TZ_SHEET_CONFIGS:
        sheet_name = cfg["sheet"]
        # Handle workbook sheets that may have trailing spaces
        ws = None
        for name in wb.sheetnames:
            if name.strip() == sheet_name.strip():
                ws = wb[name]
                break
        if ws is None:
            print(f"  Warning: sheet '{sheet_name}' not found — skipping")
            continue

        data_rows = [r for r in ws.iter_rows(values_only=True) if r[0] and r[0] != "Compound"]
        print(f"  Sheet '{sheet_name.strip()}': {len(data_rows)} rows")

        # Register this scaffold (no SMILES — to be added for Stage 3)
        sc = cfg["scaffold_code"]
        if sc not in bb_map["Scaffold"]:
            bb_map["Scaffold"][None if sc == sc else sc] = sc  # placeholder key trick below
        # Use scaffold_code as key directly (no SMILES for scaffold)
        bb_map["Scaffold"][cfg["scaffold_code"]] = cfg["scaffold_code"]

        ap = cfg["amine_prefix"]
        kp = cfg["alkyne_prefix"]
        cols = cfg["cols"]

        for row in data_rows:
            cid = str(row[0])
            m   = cfg["id_re"].match(cid)
            if not m:
                print(f"    Warning: cannot parse '{cid}' with pattern {cfg['id_re'].pattern}")
                continue

            amine_code  = f"{ap}{m.group(1)}"
            alkyne_code = f"{kp}{m.group(2)}"

            # Split concatenated SMILES
            smiles_raw = str(row[1]) if row[1] else ""
            parts      = smiles_raw.split(".")
            amine_smi  = parts[0] if len(parts) >= 1 else None
            alkyne_smi = parts[1] if len(parts) >= 2 else None

            if amine_smi  and amine_code  not in bb_map["Amine"]:
                bb_map["Amine"][amine_smi]   = amine_code
            if alkyne_smi and alkyne_code not in bb_map["Alkyne"]:
                bb_map["Alkyne"][alkyne_smi] = alkyne_code

            # Properties
            def gc(col_key):
                ci = cols.get(col_key)
                return clean_val(row[ci]) if ci is not None else None

            props = {
                "peak_pct":  gc("peak_pct"),
                "peak_norm": gc("peak_norm"),
                "rt":        gc("rt"),
                "sdr":       gc("sdr"),
                "mic":       gc("mic"),
                "tox_avg":   gc("tox_avg"),
                "tox_sd":    gc("tox_sd"),
            }
            all_compounds.append({
                "id": cid,
                "blocks": {
                    "Scaffold": cfg["scaffold_code"],
                    "Amine":    amine_code,
                    "Alkyne":   alkyne_code,
                },
                "props": props,
            })

    # Scaffold building blocks have no SMILES (stored as code→code map)
    # Rebuild bb_map["Scaffold"] so render_building_blocks gets None as SMILES key
    scaffold_entries = {}
    seen_scaffolds = set()
    for cfg in TZ_SHEET_CONFIGS:
        sc = cfg["scaffold_code"]
        if sc not in seen_scaffolds:
            scaffold_entries[sc] = cfg["scaffold_label"]
            seen_scaffolds.add(sc)

    # Produce building_blocks manually for Scaffold (no SMILES/SVG yet)
    scaffold_bbs = [
        {"code": code, "smiles": None, "name": label, "canonical_key": None, "svg": None}
        for code, label in scaffold_entries.items()
    ]

    # Render Amine and Alkyne SVGs
    print("Generating SVGs for Amine and Alkyne building blocks ...")
    amine_alkyne_map  = {"Amine": bb_map["Amine"], "Alkyne": bb_map["Alkyne"]}
    rendered          = render_building_blocks(amine_alkyne_map)

    building_blocks = {
        "Scaffold": scaffold_bbs,
        "Amine":    rendered["Amine"],
        "Alkyne":   rendered["Alkyne"],
    }

    print(f"  Scaffold: {len(scaffold_bbs)} | Amine: {len(rendered['Amine'])} | Alkyne: {len(rendered['Alkyne'])}")
    print(f"  Total compounds: {len(all_compounds)}")

    meta = LIBRARY_META[library_id]
    return {
        "id": library_id, "title": meta["title"], "description": meta["description"],
        "metal": meta["metal"], "scaffold": meta["scaffold"], "doi": meta["doi"],
        "positions":       [{"key": p["key"], "label": p["label"]} for p in TZ_POSITIONS],
        "properties":      TZ_PROPERTIES,
        "building_blocks": building_blocks,
        "compounds":       all_compounds,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# NOSB — N,O-Schiff Base Ru/Ir library (2 sheets, corrupted source SMILES columns)
# ═══════════════════════════════════════════════════════════════════════════════
#
# The source workbook's Aldehyde SMILES column (both sheets) and Aromatic SMILES
# column (RuCy sheet) have a ring-closure digit that increments with row position
# instead of staying fixed, e.g. an aldehyde meant to be constant within a
# lettered block reads "...C=C1", "...C=C2", "...C=C3" for amines 1, 2, 3 — only
# the "...C=C1" form parses. Verified by amine/aldehyde cross-checking between
# the IrCp and RuCy sheets (identical building blocks, only metal scaffold
# differs) and by RDKit parse failures on every other row. The values below are
# the one parseable row per block, taken directly from the workbook — not
# independently sourced. IrCp's Aromatic SMILES is a separate, unrelated bug
# (constant but invalid on its own — five ring positions instead of six);
# corrected to Cp* (pentamethylcyclopentadiene) per lab convention (matches
# IrCpSB's existing Cp ligand SMILES) since the source gave no parseable value
# to recover at all.

NOSB_ALDEHYDE_SMILES = {
    "A": "O=CC1=C(O)C=CC=C1",
    "B": "O=CC1=C(O)C(Br)=CC(Cl)=C1",
    "C": "O=CC1=C(O)C(OC)=CC=C1Br",
    "D": "O=CC1=C(O)C=C(N(CC)CC)C=C1",
    "E": "O=CC1=C(O)C=C(OC)C=C1",
    "F": "O=CC1=C(O)C=C(C)C=C1",
    "G": "O=CC1=C(O)C=CC([N+]([O-])=O)=C1",
    "H": "O=CC1=C(O)C=CC(Br)=C1",
    "I": "O=CC1=C(O)C=CC(C)=C1",
    "J": "O=CC1=C(O)C=CC2=C1C=CC=C2",
    "K": "O=CC1=C(O)C(CCCN2CCC3)=C2C3=C1",
}

NOSB_PROPERTIES = [
    {"key": "conversion", "label": "Conversion",          "unit": "%",   "role": "qc",       "group": None},
    {"key": "rt_target",  "label": "RT (Target)",         "unit": "min", "role": "qc",       "group": None},
    {"key": "sa_50",      "label": "S. aureus 50 µM",     "unit": "OD",  "role": "primary",  "group": "Antibacterial"},
    {"key": "sa_12",      "label": "S. aureus 12.5 µM",   "unit": "OD",  "role": "primary",  "group": "Antibacterial"},
    {"key": "ec_50",      "label": "E. coli 50 µM",       "unit": "OD",  "role": "primary",  "group": "Antibacterial"},
    {"key": "ec_100",     "label": "E. coli 100 µM",      "unit": "OD",  "role": "primary",  "group": "Antibacterial"},
    {"key": "hek_50",     "label": "HEK293T 50 µM",       "unit": "%",   "role": "primary",  "group": "Cytotoxicity"},
    {"key": "hek_sd",     "label": "HEK293T SD",          "unit": "%",   "role": "replicate","group": "Cytotoxicity"},
]

# Each entry: which sheet, the metal scaffold it represents (code/label/SMILES,
# hardcoded rather than read from the corrupted Aromatic SMILES column — see
# note above), the compound-ID pattern, and a column map per property
# (rep_cols + avg_col for {avg, reps} properties; a single col for hek_sd; None
# for assays this sheet never ran).
NOSB_SHEET_CONFIGS = [
    # Column indices are 0-based (tuple position from ws.iter_rows(values_only=True),
    # i.e. Excel column letter's 1-based position minus 1 — A=0, B=1, ... AA=26).
    {
        "sheet": "IrCp",
        "scaffold_code": "IrCp", "scaffold_label": "Ir Cp*",
        "scaffold_smiles": "CC1=C(C)C(C)=C(C)C1C",
        "id_re": re.compile(r"^IrCP_HO-SB_([A-Za-z]+)(\d+)$"),
        "cols": {
            "conversion": 4, "rt_target": 5,
            "sa_50":  ([7, 8, 9, 10],   11),
            "sa_12":  ([13, 14, 15, 16], 17),
            "ec_50":  (None, None),
            "ec_100": ([19, 20, 21, 22], 23),
            "hek_50": ([26, 27, 28, 29], 30),
            "hek_sd": 31,
        },
    },
    {
        "sheet": "RuCy",
        "scaffold_code": "RuCy", "scaffold_label": "Ru Cymene",
        "scaffold_smiles": "CC1=CC=C(C(C)C)C=C1",
        "id_re": re.compile(r"^RuCy_HO-SB_([A-Za-z]+)(\d+)$"),
        "cols": {
            "conversion": 4, "rt_target": 5,
            "sa_50":  ([7, 8, 9, 10],   11),
            "sa_12":  ([13, 14, 15, 16], 17),
            "ec_50":  ([19, 20, 21, 22], 23),
            "ec_100": ([25, 26, 27, 28], 29),
            "hek_50": ([31, 32, 33, 34], 35),
            "hek_sd": 36,
        },
    },
]


def convert_nosb(wb, library_id):
    bb_map = {"Scaffold": {}, "Aldehyde": {}, "Amine": {}}
    all_compounds = []

    for cfg in NOSB_SHEET_CONFIGS:
        ws = wb[cfg["sheet"]]
        data_rows = [r for r in ws.iter_rows(values_only=True) if r[0] and str(r[0]).strip() != "Compounds"]
        print(f"  Sheet '{cfg['sheet']}': {len(data_rows)} rows")

        bb_map["Scaffold"][cfg["scaffold_smiles"]] = cfg["scaffold_code"]

        cols = cfg["cols"]
        skipped = 0
        for row in data_rows:
            cid = str(row[0]).strip()
            m = cfg["id_re"].match(cid)
            if not m:
                print(f"    Warning: cannot parse '{cid}' with pattern {cfg['id_re'].pattern}")
                skipped += 1
                continue

            ald_code, amine_code = m.group(1), m.group(2)
            amine_smi = row[1]
            ald_smi   = NOSB_ALDEHYDE_SMILES[ald_code]

            if amine_smi and amine_code not in bb_map["Amine"]:
                bb_map["Amine"][amine_smi] = amine_code
            if ald_code not in bb_map["Aldehyde"]:
                bb_map["Aldehyde"][ald_smi] = ald_code

            props = {}
            for key in ("conversion", "rt_target"):
                props[key] = clean_val(row[cols[key]])
            for key in ("sa_50", "sa_12", "ec_50", "ec_100", "hek_50"):
                rep_cols, avg_col = cols[key]
                if avg_col is None:
                    props[key] = None
                else:
                    props[key] = {"avg": clean_val(row[avg_col]), "reps": [clean_val(row[i]) for i in rep_cols]}
            props["hek_sd"] = clean_val(row[cols["hek_sd"]]) if cols["hek_sd"] is not None else None

            all_compounds.append({
                "id": cid,
                "blocks": {
                    "Scaffold": cfg["scaffold_code"],
                    "Aldehyde": ald_code,
                    "Amine":    amine_code,
                },
                "props": props,
            })
        if skipped:
            print(f"  {skipped} rows skipped")

    print("Generating SVGs ...")
    building_blocks = render_building_blocks(bb_map)
    print(f"  Scaffold: {len(building_blocks['Scaffold'])} | "
          f"Aldehyde: {len(building_blocks['Aldehyde'])} | "
          f"Amine: {len(building_blocks['Amine'])}")
    print(f"  Total compounds: {len(all_compounds)}")

    meta = LIBRARY_META[library_id]
    return {
        "id": library_id, "title": meta["title"], "description": meta["description"],
        "metal": meta["metal"], "scaffold": meta["scaffold"], "doi": meta["doi"],
        "positions": [
            {"key": "Scaffold", "label": "Metal Scaffold"},
            {"key": "Aldehyde", "label": "Aldehyde"},
            {"key": "Amine",    "label": "Amine"},
        ],
        "properties":      NOSB_PROPERTIES,
        "building_blocks": building_blocks,
        "compounds":       all_compounds,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# MnSB — Manganese(I) Schiff-Base library (single clean sheet, MIC only)
# ═══════════════════════════════════════════════════════════════════════════════
#
# Source is Manganese420_GramPositive_MRSA_Consolidated.xlsx — already a clean,
# validated, single-sheet workbook (see WebsiteDatabase/), not a raw lab file,
# so this converter is much simpler than the others: no multi-sheet reconciling,
# no corrupted-column workarounds. Deliberately scoped to MIC only per request —
# the workbook's OD600/Screen_Conc/PublishedSI columns are intentionally ignored.

MNSB_POSITIONS = [
    {"key": "AxialLigand", "label": "Axial Ligand"},
    {"key": "Amine",       "label": "Amine"},
    {"key": "Aldehyde",    "label": "Aldehyde"},
]

MNSB_PROPERTIES = [
    {"key": "mic_um", "label": "MIC (MRSA)", "unit": "µM", "role": "primary", "group": "Antibacterial"},
]

# AxialLigand_ID (0-5) -> short code used in building blocks / compound IDs.
# Matches the published paper's own naming (Table S1).
MNSB_AXIAL_CODE = {"0": "wo", "1": "MeIm", "2": "Clo", "3": "MeBeIm", "4": "DMAP", "5": "Quin"}


def _parse_mnsb_mic(raw):
    """Returns (numeric_value_or_None, raw_text_or_None).

    Plain numbers -> (float, None).
    0 is the source's confirmed-inactive sentinel (tested, no activity found at
    any dose — NOT untested) — displayed identically to a right-censored
    ">100" result rather than literally "0.00 µM", since both mean the same
    thing in practice and 0 would otherwise sort/colour as the MOST potent
    compound, which is backwards. (User decision 2026-06-26, after noticing
    237/420 compounds showing "0.00 µM" on the site.)
    Right-censored (">100") -> (100.0, ">100").
    A dilution-bracket range ("50-25") -> (50.0, "50-25") — the upper/first
    bound, so the numeric value is always the conservative (less potent) end.
    """
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)):
        if float(raw) == 0:
            return 100.0, ">100"
        return float(raw), None
    s = str(raw).strip()
    if s == "":
        return None, None
    if s.startswith(">"):
        try:
            return float(s[1:].strip()), s
        except ValueError:
            return None, s
    if "-" in s:
        try:
            return float(s.split("-")[0].strip()), s
        except ValueError:
            return None, s
    try:
        return float(s), None
    except ValueError:
        return None, s


def convert_mnsb(wb, library_id):
    ws = wb["Compounds"]
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    col = {name: i for i, name in enumerate(header)}
    print(f"  {len(data_rows)} compound rows")

    bb_map = {"AxialLigand": {}, "Amine": {}, "Aldehyde": {}}
    compounds = []
    for row in data_rows:
        axial_id  = str(row[col["AxialLigand_ID"]])
        amine_id  = str(row[col["Amine_ID"]])
        ald_id    = str(row[col["Aldehyde_ID"]])
        axial_smi = row[col["AxialLigand_SMILES"]]
        amine_smi = row[col["Amine_SMILES"]]
        ald_smi   = row[col["Aldehyde_SMILES"]]

        axial_code = MNSB_AXIAL_CODE[axial_id]
        amine_code = f"Am{amine_id}"
        ald_code   = f"Al{ald_id}"

        if axial_smi not in bb_map["AxialLigand"]:
            bb_map["AxialLigand"][axial_smi] = axial_code  # None SMILES for "wo" — fine, same as TzLib's null scaffolds
        if amine_smi and amine_smi not in bb_map["Amine"]:
            bb_map["Amine"][amine_smi] = amine_code
        if ald_smi and ald_smi not in bb_map["Aldehyde"]:
            bb_map["Aldehyde"][ald_smi] = ald_code

        mic_val, mic_raw = _parse_mnsb_mic(row[col["MIC_MRSA_uM"]])
        compound = {
            "id": f"{axial_code}_{amine_code}_{ald_code}",
            "blocks": {"AxialLigand": axial_code, "Amine": amine_code, "Aldehyde": ald_code},
            "props": {"mic_um": mic_val},
        }
        if mic_raw is not None:
            compound["mic_raw"] = mic_raw
        compounds.append(compound)

    print("Generating SVGs ...")
    building_blocks = render_building_blocks(bb_map)
    print(f"  AxialLigand: {len(building_blocks['AxialLigand'])} | "
          f"Amine: {len(building_blocks['Amine'])} | "
          f"Aldehyde: {len(building_blocks['Aldehyde'])}")
    print(f"  Total compounds: {len(compounds)}")

    meta = LIBRARY_META[library_id]
    return {
        "id": library_id, "title": meta["title"], "description": meta["description"],
        "metal": meta["metal"], "scaffold": meta["scaffold"], "doi": meta["doi"],
        "positions":       MNSB_POSITIONS,
        "properties":      MNSB_PROPERTIES,
        "building_blocks": building_blocks,
        "compounds":       compounds,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# IrCN_Click / IrCN_Schiff — bis-cyclometalated Ir(III) click and Schiff-base
# libraries. Both store assembled complex SMILES in a single column; individual
# BB SMILES are extracted by splitting on '.' (format: [Ir].ppy.ppy.bb3.bb4).
# ═══════════════════════════════════════════════════════════════════════════════

IRCN_CLICK_POSITIONS = [
    {"key": "Scaffold", "label": "Metal Scaffold"},
    {"key": "Amine",    "label": "Amine"},
    {"key": "Alkyne",   "label": "Alkyne"},
]

IRCN_CLICK_PROPERTIES = [
    {"key": "peak_pct", "label": "Conversion",      "unit": "%",   "role": "qc",      "group": None},
    {"key": "rt",       "label": "Retention Time",   "unit": "min", "role": "qc",      "group": None},
    {"key": "mic",      "label": "MIC (S. aureus)",  "unit": "µM",  "role": "primary", "group": "Antibacterial"},
]

IRCN_SCHIFF_POSITIONS = [
    {"key": "Scaffold",  "label": "Metal Scaffold"},
    {"key": "Aldehyde",  "label": "Aldehyde"},
    {"key": "Amine",     "label": "Amine"},
]

IRCN_SCHIFF_PROPERTIES = [
    {"key": "peak_pct", "label": "Conversion",      "unit": "%",   "role": "qc",      "group": None},
    {"key": "rt",       "label": "Retention Time",   "unit": "min", "role": "qc",      "group": None},
    {"key": "mic",      "label": "MIC (S. aureus)",  "unit": "µM",  "role": "primary", "group": "Antibacterial"},
]

_IRCN_CLICK_ID_RE  = re.compile(r'^(IrCN[123])_(M\d+)(Y\d+)$')
_IRCN_SCHIFF_ID_RE = re.compile(r'^(IrCN[123])_(P\d+)(A\d+)$')


def _ircn_parse_assembled(cid, assembled, click=True):
    """Split '[Ir].ppy.ppy.bb3.bb4' into (scaffold_smi, bb3_smi, bb4_smi).
    Returns None on unexpected fragment count."""
    frags = assembled.split(".")
    if len(frags) < 5:
        print(f"  Warning: {len(frags)} fragments (expected ≥5) for '{cid}' — skipping")
        return None
    return frags[1], frags[3], frags[4]


def convert_ircn_click(wb, library_id):
    ws = wb["IrCN click compounds"]
    data_rows = [r for r in ws.iter_rows(values_only=True)
                 if r[0] and r[0] != "File_name"]
    print(f"  {len(data_rows)} compound rows")

    bb_scaffold = {}   # ppy_smi -> scaffold_code (first seen SMILES per scaffold code)
    scaffold_seen = {} # scaffold_code -> ppy_smi (to avoid duplicate key insertions)
    bb_amine     = {}  # amine_smi -> amine_code
    bb_alkyne    = {}  # alkyne_smi -> alkyne_code
    compounds = []
    skipped = 0

    for row in data_rows:
        cid = str(row[0])
        m = _IRCN_CLICK_ID_RE.match(cid)
        if not m:
            print(f"  Warning: cannot parse '{cid}' — skipping")
            skipped += 1
            continue

        scaffold_code = m.group(1)   # IrCN1 / IrCN2 / IrCN3
        amine_code    = m.group(2)   # M1 … M24
        alkyne_code   = m.group(3)   # Y1 … Y4

        parsed = _ircn_parse_assembled(cid, str(row[1]) if row[1] else "")
        if parsed is None:
            skipped += 1
            continue
        ppy_smi, amine_smi, alkyne_smi = parsed

        if scaffold_code not in scaffold_seen:
            scaffold_seen[scaffold_code] = ppy_smi
            bb_scaffold[ppy_smi] = scaffold_code
        if amine_smi and amine_smi not in bb_amine:
            bb_amine[amine_smi] = amine_code
        if alkyne_smi and alkyne_smi not in bb_alkyne:
            bb_alkyne[alkyne_smi] = alkyne_code

        compounds.append({
            "id": cid,
            "blocks": {"Scaffold": scaffold_code, "Amine": amine_code, "Alkyne": alkyne_code},
            "props": {
                "peak_pct": clean_val(row[2]),
                "rt":       clean_val(row[3]),
                "mic":      clean_val(row[4]),
            },
        })

    if skipped:
        print(f"  {skipped} rows skipped")
    print("Generating SVGs ...")
    building_blocks = render_building_blocks({
        "Scaffold": bb_scaffold,
        "Amine":    bb_amine,
        "Alkyne":   bb_alkyne,
    })
    print(f"  Scaffold: {len(building_blocks['Scaffold'])} | "
          f"Amine: {len(building_blocks['Amine'])} | "
          f"Alkyne: {len(building_blocks['Alkyne'])}")
    print(f"  Total compounds: {len(compounds)}")

    meta = LIBRARY_META[library_id]
    return {
        "id": library_id, "title": meta["title"], "description": meta["description"],
        "metal": meta["metal"], "scaffold": meta["scaffold"], "doi": meta["doi"],
        "positions":       IRCN_CLICK_POSITIONS,
        "properties":      IRCN_CLICK_PROPERTIES,
        "building_blocks": building_blocks,
        "compounds":       compounds,
    }


def convert_ircn_schiff(wb, library_id):
    ws = wb["IrCN Schiff compounds"]
    data_rows = [r for r in ws.iter_rows(values_only=True)
                 if r[0] and r[0] != "New name"]
    print(f"  {len(data_rows)} compound rows")

    bb_scaffold    = {}  # ppy_smi -> scaffold_code
    scaffold_seen  = {}  # scaffold_code -> ppy_smi
    bb_aldehyde    = {}  # ald_smi -> aldehyde_code
    bb_amine       = {}  # amine_smi -> amine_code
    compounds = []
    skipped = 0

    for row in data_rows:
        cid = str(row[0])
        m = _IRCN_SCHIFF_ID_RE.match(cid)
        if not m:
            print(f"  Warning: cannot parse '{cid}' — skipping")
            skipped += 1
            continue

        scaffold_code  = m.group(1)  # IrCN1 / IrCN2 / IrCN3
        aldehyde_code  = m.group(2)  # P1 … P8
        amine_code     = m.group(3)  # A1 … A11

        # col[1] is blank; assembled SMILES is at col[2]
        parsed = _ircn_parse_assembled(cid, str(row[2]) if row[2] else "")
        if parsed is None:
            skipped += 1
            continue
        ppy_smi, aldehyde_smi, amine_smi = parsed

        if scaffold_code not in scaffold_seen:
            scaffold_seen[scaffold_code] = ppy_smi
            bb_scaffold[ppy_smi] = scaffold_code
        if aldehyde_smi and aldehyde_smi not in bb_aldehyde:
            bb_aldehyde[aldehyde_smi] = aldehyde_code
        if amine_smi and amine_smi not in bb_amine:
            bb_amine[amine_smi] = amine_code

        compounds.append({
            "id": cid,
            "blocks": {"Scaffold": scaffold_code, "Aldehyde": aldehyde_code, "Amine": amine_code},
            "props": {
                "peak_pct": clean_val(row[3]),
                "rt":       clean_val(row[4]),
                "mic":      clean_val(row[5]),
            },
        })

    if skipped:
        print(f"  {skipped} rows skipped")
    print("Generating SVGs ...")
    building_blocks = render_building_blocks({
        "Scaffold": bb_scaffold,
        "Aldehyde": bb_aldehyde,
        "Amine":    bb_amine,
    })
    print(f"  Scaffold: {len(building_blocks['Scaffold'])} | "
          f"Aldehyde: {len(building_blocks['Aldehyde'])} | "
          f"Amine: {len(building_blocks['Amine'])}")
    print(f"  Total compounds: {len(compounds)}")

    meta = LIBRARY_META[library_id]
    return {
        "id": library_id, "title": meta["title"], "description": meta["description"],
        "metal": meta["metal"], "scaffold": meta["scaffold"], "doi": meta["doi"],
        "positions":       IRCN_SCHIFF_POSITIONS,
        "properties":      IRCN_SCHIFF_PROPERTIES,
        "building_blocks": building_blocks,
        "compounds":       compounds,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════════════════════

CONVERTERS = {
    "IrCpSB":     convert_ircpsb,
    "TzLib":      convert_tzlib,
    "NOSB":       convert_nosb,
    "MnSB":       convert_mnsb,
    "IrCN_Click": convert_ircn_click,
    "IrCN_Schiff":convert_ircn_schiff,
}


def convert(excel_path, output_dir, library_id):
    if library_id not in LIBRARY_META:
        sys.exit(f"Unknown library ID '{library_id}'. Valid IDs: {list(LIBRARY_META)}")

    print(f"Reading {excel_path} ...")
    wb      = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    library = CONVERTERS[library_id](wb, library_id)
    write_outputs(library, output_dir)
    print("Done.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert Frei Lab Excel → JSON")
    parser.add_argument("excel",  help="Path to the source Excel file")
    parser.add_argument("output", help="Output directory (e.g. ../public/data)")
    parser.add_argument("--library-id", required=True, help=f"One of: {list(LIBRARY_META)}")
    args = parser.parse_args()
    convert(args.excel, args.output, args.library_id)
